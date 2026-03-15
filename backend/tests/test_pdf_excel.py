"""
Test PDF Template Selection and Excel Import/Export functionality for QuoteMaster Pro
Features tested:
- PDF templates endpoint
- PDF template selection and user update
- Excel export for products
- Excel import for products
- PDF generation with selected template
"""

import pytest
import requests
import os
import io
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL')

class TestSetup:
    """Setup fixtures for authentication"""
    
    @pytest.fixture(scope='class')
    def auth_token(self):
        """Register/login a test user and return auth token"""
        # Try to register first
        register_data = {
            "email": "test_pdf_excel@example.com",
            "password": "TestPass123!",
            "company_name": "TEST PDF Excel Company"
        }
        
        response = requests.post(f"{BASE_URL}/api/auth/register", json=register_data)
        if response.status_code == 200:
            return response.json()["token"]
        
        # If already exists, try login
        login_data = {
            "email": "test_pdf_excel@example.com",
            "password": "TestPass123!"
        }
        response = requests.post(f"{BASE_URL}/api/auth/login", json=login_data)
        if response.status_code == 200:
            return response.json()["token"]
        
        pytest.skip("Could not authenticate test user")
    
    @pytest.fixture(scope='class')
    def auth_headers(self, auth_token):
        """Return headers with auth token"""
        return {
            "Authorization": f"Bearer {auth_token}",
            "Content-Type": "application/json"
        }


class TestPDFTemplates(TestSetup):
    """Test PDF Template Selection Feature"""
    
    def test_get_pdf_templates(self, auth_headers):
        """Test GET /api/pdf-templates returns available templates"""
        response = requests.get(f"{BASE_URL}/api/pdf-templates", headers=auth_headers)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        templates = response.json()
        assert isinstance(templates, list), "Templates should be a list"
        assert len(templates) >= 5, f"Expected at least 5 templates, got {len(templates)}"
        
        # Verify template structure
        template_ids = []
        for template in templates:
            assert "id" in template, "Template should have 'id'"
            assert "name" in template, "Template should have 'name'"
            assert "primary" in template, "Template should have 'primary' color"
            assert "accent" in template, "Template should have 'accent' color"
            template_ids.append(template["id"])
        
        # Verify expected templates exist
        expected_ids = ["classic", "modern", "professional", "elegant", "ocean"]
        for expected_id in expected_ids:
            assert expected_id in template_ids, f"Template '{expected_id}' not found"
        
        print(f"✓ Found {len(templates)} PDF templates: {template_ids}")
    
    def test_select_pdf_template(self, auth_headers):
        """Test updating user's PDF template selection"""
        # Select 'modern' template
        update_data = {"pdf_template": "modern"}
        response = requests.put(f"{BASE_URL}/api/auth/profile", json=update_data, headers=auth_headers)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        user_data = response.json()
        assert user_data.get("pdf_template") == "modern", f"Expected 'modern' template, got {user_data.get('pdf_template')}"
        
        print("✓ PDF template selection updated to 'modern'")
    
    def test_select_all_templates(self, auth_headers):
        """Test selecting each available template"""
        templates_to_test = ["classic", "modern", "professional", "elegant", "ocean"]
        
        for template_id in templates_to_test:
            update_data = {"pdf_template": template_id}
            response = requests.put(f"{BASE_URL}/api/auth/profile", json=update_data, headers=auth_headers)
            
            assert response.status_code == 200, f"Failed to select template '{template_id}': {response.text}"
            assert response.json().get("pdf_template") == template_id, f"Template not set to '{template_id}'"
        
        # Set back to classic
        requests.put(f"{BASE_URL}/api/auth/profile", json={"pdf_template": "classic"}, headers=auth_headers)
        print(f"✓ All {len(templates_to_test)} templates can be selected")


class TestExcelExport(TestSetup):
    """Test Excel Export Feature"""
    
    def test_create_test_products(self, auth_headers):
        """Create test products for export testing"""
        test_products = [
            {
                "name": "TEST_EXPORT_Product_1",
                "description": "Test product for export",
                "sku": "TEST-EXP-001",
                "unit": "adet",
                "unit_price": 100.00,
                "vat_rate": 20
            },
            {
                "name": "TEST_EXPORT_Product_2",
                "description": "Another test product",
                "sku": "TEST-EXP-002",
                "unit": "kg",
                "unit_price": 250.50,
                "vat_rate": 10
            }
        ]
        
        created_ids = []
        for product in test_products:
            response = requests.post(f"{BASE_URL}/api/products", json=product, headers=auth_headers)
            assert response.status_code == 200, f"Failed to create product: {response.text}"
            created_ids.append(response.json()["id"])
        
        print(f"✓ Created {len(created_ids)} test products for export")
        return created_ids
    
    def test_export_products_excel(self, auth_headers):
        """Test GET /api/products/export/excel"""
        # Ensure we have at least one product
        self.test_create_test_products(auth_headers)
        
        response = requests.get(f"{BASE_URL}/api/products/export/excel", headers=auth_headers)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", ""), \
            f"Expected Excel content type, got {response.headers.get('Content-Type')}"
        
        # Verify it's a valid Excel file
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Verify headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat (₺)", "KDV Oranı (%)"]
        assert headers == expected_headers, f"Headers mismatch: {headers} != {expected_headers}"
        
        # Verify data rows exist
        row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if any(_))
        assert row_count > 0, "Excel should have product data rows"
        
        print(f"✓ Excel export successful: {row_count} products exported")


class TestExcelImport(TestSetup):
    """Test Excel Import Feature"""
    
    def test_import_products_excel(self, auth_headers):
        """Test POST /api/products/import/excel"""
        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Ürünler"
        
        # Headers
        headers = ["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat (₺)", "KDV Oranı (%)"]
        ws.append(headers)
        
        # Test products to import
        ws.append(["TEST-IMP-001", "TEST_IMPORT_Product_1", "Imported product 1", "adet", 150.00, 20])
        ws.append(["TEST-IMP-002", "TEST_IMPORT_Product_2", "Imported product 2", "m", 300.00, 10])
        ws.append(["TEST-IMP-003", "TEST_IMPORT_Product_3", "Imported product 3", "kg", 75.50, 0])
        
        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload file
        files = {"file": ("test_import.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers_multipart = {"Authorization": auth_headers["Authorization"]}
        
        response = requests.post(f"{BASE_URL}/api/products/import/excel", files=files, headers=headers_multipart)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        assert "imported" in result, "Response should have 'imported' count"
        assert "updated" in result, "Response should have 'updated' count"
        
        total_processed = result["imported"] + result["updated"]
        assert total_processed == 3, f"Expected 3 products processed, got {total_processed}"
        
        print(f"✓ Excel import successful: {result['imported']} imported, {result['updated']} updated")
    
    def test_import_updates_existing_products(self, auth_headers):
        """Test that importing with same SKU updates existing product"""
        # First import
        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat (₺)", "KDV Oranı (%)"])
        ws.append(["TEST-UPDATE-001", "TEST_Original_Name", "Original description", "adet", 100.00, 20])
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {"file": ("first_import.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers_multipart = {"Authorization": auth_headers["Authorization"]}
        
        response = requests.post(f"{BASE_URL}/api/products/import/excel", files=files, headers=headers_multipart)
        assert response.status_code == 200
        first_result = response.json()
        
        # Second import with same SKU but different name
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat (₺)", "KDV Oranı (%)"])
        ws2.append(["TEST-UPDATE-001", "TEST_Updated_Name", "Updated description", "adet", 200.00, 10])
        
        excel_buffer2 = io.BytesIO()
        wb2.save(excel_buffer2)
        excel_buffer2.seek(0)
        
        files2 = {"file": ("second_import.xlsx", excel_buffer2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        response2 = requests.post(f"{BASE_URL}/api/products/import/excel", files=files2, headers=headers_multipart)
        assert response2.status_code == 200
        second_result = response2.json()
        
        # Should update instead of import new
        assert second_result["updated"] >= 1, f"Expected at least 1 update, got {second_result['updated']}"
        
        print(f"✓ Excel import updates existing products by SKU")
    
    def test_verify_imported_products(self, auth_headers):
        """Verify imported products exist in database"""
        response = requests.get(f"{BASE_URL}/api/products", headers=auth_headers)
        assert response.status_code == 200
        
        products = response.json()
        product_skus = [p.get("sku") for p in products if p.get("sku")]
        
        # Check if our test imports exist
        expected_skus = ["TEST-IMP-001", "TEST-IMP-002", "TEST-IMP-003"]
        for sku in expected_skus:
            assert sku in product_skus, f"Imported product with SKU '{sku}' not found"
        
        print(f"✓ All imported products verified in database")


class TestPDFGenerationWithTemplate(TestSetup):
    """Test PDF Generation with Selected Template"""
    
    @pytest.fixture(scope='class')
    def test_product(self, auth_headers):
        """Create a test product for quote"""
        product_data = {
            "name": "TEST_PDF_Product",
            "description": "Product for PDF test",
            "sku": "TEST-PDF-001",
            "unit": "adet",
            "unit_price": 500.00,
            "vat_rate": 20
        }
        response = requests.post(f"{BASE_URL}/api/products", json=product_data, headers=auth_headers)
        if response.status_code == 200:
            return response.json()
        pytest.skip("Could not create test product")
    
    @pytest.fixture(scope='class')
    def test_quote(self, auth_headers, test_product):
        """Create a test quote for PDF generation"""
        quote_data = {
            "customer_name": "TEST PDF Customer",
            "customer_email": "pdftest@example.com",
            "customer_phone": "0555 555 55 55",
            "customer_address": "Test Address, Istanbul",
            "customer_tax_number": "1234567890",
            "items": [
                {
                    "product_id": test_product["id"],
                    "product_name": test_product["name"],
                    "unit": test_product["unit"],
                    "quantity": 5,
                    "unit_price": test_product["unit_price"],
                    "vat_rate": test_product["vat_rate"],
                    "discount_percent": 10
                }
            ],
            "validity_days": 30,
            "notes": "Test quote for PDF generation",
            "include_vat": True
        }
        response = requests.post(f"{BASE_URL}/api/quotes", json=quote_data, headers=auth_headers)
        if response.status_code == 200:
            return response.json()
        pytest.skip(f"Could not create test quote: {response.text}")
    
    def test_generate_pdf_with_classic_template(self, auth_headers, test_quote):
        """Test PDF generation with classic template"""
        # Set template to classic
        requests.put(f"{BASE_URL}/api/auth/profile", json={"pdf_template": "classic"}, headers=auth_headers)
        
        # Generate PDF
        response = requests.get(f"{BASE_URL}/api/quotes/{test_quote['id']}/pdf", headers=auth_headers)
        
        assert response.status_code == 200, f"PDF generation failed: {response.text}"
        assert "application/pdf" in response.headers.get("Content-Type", ""), \
            f"Expected PDF content type, got {response.headers.get('Content-Type')}"
        
        # Verify PDF is not empty
        assert len(response.content) > 1000, f"PDF seems too small: {len(response.content)} bytes"
        
        # Verify PDF header
        assert response.content[:4] == b'%PDF', "Invalid PDF format"
        
        print(f"✓ PDF generated with classic template ({len(response.content)} bytes)")
    
    def test_generate_pdf_with_modern_template(self, auth_headers, test_quote):
        """Test PDF generation with modern template"""
        # Set template to modern
        requests.put(f"{BASE_URL}/api/auth/profile", json={"pdf_template": "modern"}, headers=auth_headers)
        
        # Generate PDF
        response = requests.get(f"{BASE_URL}/api/quotes/{test_quote['id']}/pdf", headers=auth_headers)
        
        assert response.status_code == 200, f"PDF generation failed: {response.text}"
        assert response.content[:4] == b'%PDF', "Invalid PDF format"
        
        print(f"✓ PDF generated with modern template ({len(response.content)} bytes)")
    
    def test_generate_pdf_with_all_templates(self, auth_headers, test_quote):
        """Test PDF generation with each available template"""
        templates = ["classic", "modern", "professional", "elegant", "ocean"]
        
        for template_id in templates:
            # Set template
            requests.put(f"{BASE_URL}/api/auth/profile", json={"pdf_template": template_id}, headers=auth_headers)
            
            # Generate PDF
            response = requests.get(f"{BASE_URL}/api/quotes/{test_quote['id']}/pdf", headers=auth_headers)
            
            assert response.status_code == 200, f"PDF generation with '{template_id}' template failed"
            assert response.content[:4] == b'%PDF', f"Invalid PDF format for '{template_id}' template"
        
        print(f"✓ PDF generation works with all {len(templates)} templates")


class TestCleanup(TestSetup):
    """Cleanup test data"""
    
    def test_cleanup_test_products(self, auth_headers):
        """Delete test products created during testing"""
        response = requests.get(f"{BASE_URL}/api/products", headers=auth_headers)
        if response.status_code != 200:
            return
        
        products = response.json()
        deleted_count = 0
        
        for product in products:
            if product.get("sku", "").startswith("TEST-") or product.get("name", "").startswith("TEST_"):
                requests.delete(f"{BASE_URL}/api/products/{product['id']}", headers=auth_headers)
                deleted_count += 1
        
        print(f"✓ Cleaned up {deleted_count} test products")
    
    def test_cleanup_test_quotes(self, auth_headers):
        """Delete test quotes created during testing"""
        response = requests.get(f"{BASE_URL}/api/quotes", headers=auth_headers)
        if response.status_code != 200:
            return
        
        quotes = response.json()
        deleted_count = 0
        
        for quote in quotes:
            if quote.get("customer_name", "").startswith("TEST"):
                requests.delete(f"{BASE_URL}/api/quotes/{quote['id']}", headers=auth_headers)
                deleted_count += 1
        
        print(f"✓ Cleaned up {deleted_count} test quotes")
