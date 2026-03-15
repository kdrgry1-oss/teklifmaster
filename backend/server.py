from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Query, BackgroundTasks, Request
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, HTMLResponse
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
import hashlib
import hmac
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import jwt
import bcrypt
from io import BytesIO
import base64
import json
import asyncio
import secrets
import html
from cryptography.fernet import Fernet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tempfile
import aiofiles
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Optional: Resend for email (only if configured)
try:
    import resend
    RESEND_AVAILABLE = True
except ImportError:
    RESEND_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', secrets.token_urlsafe(32))
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Encryption key for sensitive data (generate if not exists)
ENCRYPTION_KEY = os.environ.get('ENCRYPTION_KEY', Fernet.generate_key().decode())
fernet = Fernet(ENCRYPTION_KEY.encode() if isinstance(ENCRYPTION_KEY, str) else ENCRYPTION_KEY)

# Iyzico Settings
IYZICO_API_KEY = os.environ.get('IYZICO_API_KEY', '')
IYZICO_SECRET_KEY = os.environ.get('IYZICO_SECRET_KEY', '')
IYZICO_BASE_URL = os.environ.get('IYZICO_BASE_URL', 'api.iyzipay.com')

# Resend Email Settings (optional)
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'info@teklifmaster.com')
if RESEND_AVAILABLE and RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# SMTP Settings for Hostinger
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.hostinger.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '465'))
SMTP_USER = os.environ.get('SMTP_USER', 'info@teklifmaster.com')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')

# Admin email
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'kdrgry@gmail.com')

# ============== SECURITY UTILITIES ==============

# Rate limiting storage (in-memory, use Redis in production)
rate_limit_storage: Dict[str, List[datetime]] = defaultdict(list)
blocked_ips: Dict[str, datetime] = {}

# Security constants
MAX_REQUESTS_PER_MINUTE = 60
MAX_LOGIN_ATTEMPTS = 5
LOGIN_BLOCK_DURATION = timedelta(minutes=15)
SUSPICIOUS_PATTERNS = [
    r'<script[^>]*>',
    r'javascript:',
    r'on\w+\s*=',
    r'data:text/html',
    r'vbscript:',
    r'expression\s*\(',
]

def sanitize_input(value: str) -> str:
    """Sanitize input to prevent XSS and injection attacks"""
    if not value:
        return value
    # HTML escape
    sanitized = html.escape(value)
    # Check for suspicious patterns
    for pattern in SUSPICIOUS_PATTERNS:
        if re.search(pattern, value, re.IGNORECASE):
            logger.warning(f"Suspicious pattern detected: {pattern}")
            sanitized = re.sub(pattern, '', sanitized, flags=re.IGNORECASE)
    return sanitized

def sanitize_dict(data: dict) -> dict:
    """Recursively sanitize all string values in a dictionary"""
    sanitized = {}
    for key, value in data.items():
        if isinstance(value, str):
            sanitized[key] = sanitize_input(value)
        elif isinstance(value, dict):
            sanitized[key] = sanitize_dict(value)
        elif isinstance(value, list):
            sanitized[key] = [
                sanitize_input(v) if isinstance(v, str) else 
                sanitize_dict(v) if isinstance(v, dict) else v 
                for v in value
            ]
        else:
            sanitized[key] = value
    return sanitized

def encrypt_sensitive_data(data: str) -> str:
    """Encrypt sensitive data"""
    return fernet.encrypt(data.encode()).decode()

def decrypt_sensitive_data(encrypted_data: str) -> str:
    """Decrypt sensitive data"""
    try:
        return fernet.decrypt(encrypted_data.encode()).decode()
    except Exception:
        return encrypted_data  # Return as-is if decryption fails

def hash_sensitive_field(value: str) -> str:
    """Create a hash for sensitive fields (for searching encrypted data)"""
    return hashlib.sha256(value.encode()).hexdigest()[:16]

def check_rate_limit(ip: str, limit: int = MAX_REQUESTS_PER_MINUTE) -> bool:
    """Check if IP has exceeded rate limit"""
    now = datetime.now(timezone.utc)
    minute_ago = now - timedelta(minutes=1)
    
    # Clean old entries
    rate_limit_storage[ip] = [t for t in rate_limit_storage[ip] if t > minute_ago]
    
    if len(rate_limit_storage[ip]) >= limit:
        return False
    
    rate_limit_storage[ip].append(now)
    return True

def is_ip_blocked(ip: str) -> bool:
    """Check if IP is blocked"""
    if ip in blocked_ips:
        if blocked_ips[ip] > datetime.now(timezone.utc):
            return True
        else:
            del blocked_ips[ip]
    return False

def block_ip(ip: str, duration: timedelta = LOGIN_BLOCK_DURATION):
    """Block an IP address"""
    blocked_ips[ip] = datetime.now(timezone.utc) + duration
    logger.warning(f"IP blocked: {ip}")

# Login attempt tracking
login_attempts: Dict[str, List[datetime]] = defaultdict(list)

def check_login_attempts(email: str, ip: str) -> bool:
    """Check if login attempts exceeded"""
    key = f"{email}:{ip}"
    now = datetime.now(timezone.utc)
    window = now - LOGIN_BLOCK_DURATION
    
    # Clean old attempts
    login_attempts[key] = [t for t in login_attempts[key] if t > window]
    
    if len(login_attempts[key]) >= MAX_LOGIN_ATTEMPTS:
        return False
    return True

def record_login_attempt(email: str, ip: str):
    """Record a failed login attempt"""
    key = f"{email}:{ip}"
    login_attempts[key].append(datetime.now(timezone.utc))

# ============== SECURITY MIDDLEWARE ==============

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """Add security headers to all responses"""
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        
        # Security headers
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        
        # Remove server header
        if "server" in response.headers:
            del response.headers["server"]
        
        return response

class RateLimitMiddleware(BaseHTTPMiddleware):
    """Rate limiting middleware"""
    async def dispatch(self, request: Request, call_next):
        client_ip = request.client.host if request.client else "unknown"
        
        # Check if IP is blocked
        if is_ip_blocked(client_ip):
            logger.warning(f"Blocked IP attempted access: {client_ip}")
            return JSONResponse(
                status_code=403,
                content={"detail": "Erişiminiz geçici olarak engellenmiştir"}
            )
        
        # Check rate limit
        if not check_rate_limit(client_ip):
            logger.warning(f"Rate limit exceeded for IP: {client_ip}")
            return JSONResponse(
                status_code=429,
                content={"detail": "Çok fazla istek. Lütfen bir dakika bekleyin."}
            )
        
        return await call_next(request)

class AuditLogMiddleware(BaseHTTPMiddleware):
    """Audit logging for sensitive operations"""
    SENSITIVE_PATHS = ['/api/auth/', '/api/quotes/', '/api/customers/', '/api/products/']
    
    async def dispatch(self, request: Request, call_next):
        # Log sensitive operations
        path = request.url.path
        method = request.method
        client_ip = request.client.host if request.client else "unknown"
        
        is_sensitive = any(path.startswith(p) for p in self.SENSITIVE_PATHS)
        
        if is_sensitive and method in ['POST', 'PUT', 'DELETE']:
            logger.info(f"AUDIT: {method} {path} from {client_ip}")
        
        response = await call_next(request)
        
        # Log failed auth attempts
        if path.startswith('/api/auth/login') and response.status_code == 401:
            logger.warning(f"AUDIT: Failed login attempt from {client_ip}")
        
        return response

# Create the main app
app = FastAPI(
    title="TeklifMaster API",
    docs_url=None,  # Disable Swagger UI in production
    redoc_url=None,  # Disable ReDoc in production
    openapi_url=None  # Disable OpenAPI schema in production
)

# Add security middleware (order matters - first added = last executed)
app.add_middleware(AuditLogMiddleware)
app.add_middleware(RateLimitMiddleware)
app.add_middleware(SecurityHeadersMiddleware)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Upload directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# ============== MODELS WITH VALIDATION ==============

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    company_name: str
    
    @field_validator('password')
    @classmethod
    def validate_password(cls, v):
        if len(v) < 6:
            raise ValueError('Şifre en az 6 karakter olmalıdır')
        if len(v) > 128:
            raise ValueError('Şifre çok uzun')
        return v
    
    @field_validator('company_name')
    @classmethod
    def validate_company_name(cls, v):
        v = sanitize_input(v.strip())
        if len(v) < 2:
            raise ValueError('Şirket adı en az 2 karakter olmalıdır')
        if len(v) > 200:
            raise ValueError('Şirket adı çok uzun')
        return v

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    company_name: str
    company_logo: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_tax_number: Optional[str] = None
    pdf_template: str = "classic"
    trial_end_date: Optional[str] = None
    subscription_status: str = "trial"
    is_admin: bool = False
    created_at: str

class UserUpdate(BaseModel):
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_tax_number: Optional[str] = None
    pdf_template: Optional[str] = None
    # PDF Ayarları
    pdf_show_images: Optional[bool] = None
    pdf_image_size: Optional[str] = None  # small, medium, large
    pdf_description_length: Optional[str] = None  # full, short, hidden

class ProductCreate(BaseModel):
    name: str
    description: Optional[str] = None
    unit: str = "adet"
    unit_price: float
    currency: str = "TRY"
    vat_rate: float = 20.0
    image_url: Optional[str] = None
    sku: Optional[str] = None
    
    @field_validator('name')
    @classmethod
    def validate_name(cls, v):
        v = sanitize_input(v.strip())
        if len(v) < 1 or len(v) > 200:
            raise ValueError('Ürün adı 1-200 karakter arası olmalıdır')
        return v
    
    @field_validator('description')
    @classmethod
    def validate_description(cls, v):
        if v:
            v = sanitize_input(v.strip())
            if len(v) > 2000:
                raise ValueError('Açıklama çok uzun')
        return v
    
    @field_validator('unit_price')
    @classmethod
    def validate_price(cls, v):
        if v < 0 or v > 999999999:
            raise ValueError('Geçersiz fiyat')
        return v

class ProductResponse(BaseModel):
    id: str
    user_id: str
    name: str
    description: Optional[str] = None
    unit: str
    unit_price: float
    currency: str = "TRY"
    vat_rate: float
    image_url: Optional[str] = None
    sku: Optional[str] = None
    created_at: str
    updated_at: str

class BankAccountCreate(BaseModel):
    bank_name: str
    iban: str
    account_holder: Optional[str] = None
    currency: str = "TRY"
    account_type: str = "checking"  # checking, savings, foreign
    
    @field_validator('bank_name')
    @classmethod
    def validate_bank_name(cls, v):
        v = sanitize_input(v.strip())
        if len(v) < 2 or len(v) > 100:
            raise ValueError('Banka adı geçersiz')
        return v
    
    @field_validator('iban')
    @classmethod
    def validate_iban(cls, v):
        # Remove spaces and validate
        v = v.replace(' ', '').upper()
        if not v.startswith('TR'):
            v = 'TR' + v
        if len(v) != 26:
            raise ValueError('IBAN 26 karakter olmalıdır')
        if not re.match(r'^TR\d{24}$', v):
            raise ValueError('Geçersiz IBAN formatı')
        return v
    
    @field_validator('account_holder')
    @classmethod
    def validate_account_holder(cls, v):
        if v:
            v = sanitize_input(v.strip())
            if len(v) > 100:
                raise ValueError('Hesap sahibi adı çok uzun')
        return v

class BankAccountResponse(BaseModel):
    id: str
    user_id: str
    bank_name: str
    iban: str
    account_holder: Optional[str] = None
    currency: str = "TRY"
    account_type: str = "checking"
    created_at: str

class QuoteItemCreate(BaseModel):
    product_id: str
    product_name: str
    unit: str
    quantity: float
    unit_price: float
    vat_rate: float
    discount_percent: float = 0.0

class QuoteCreate(BaseModel):
    quote_name: str
    customer_id: Optional[str] = None
    customer_name: str
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    customer_tax_number: Optional[str] = None
    items: List[QuoteItemCreate]
    bank_account_ids: List[str] = []
    validity_days: int = 30
    notes: Optional[str] = None
    include_vat: bool = True
    # Genel İskonto
    general_discount_type: Optional[str] = None  # percent, amount, None
    general_discount_value: float = 0.0

class QuoteUpdate(BaseModel):
    quote_name: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    customer_tax_number: Optional[str] = None
    items: Optional[List[QuoteItemCreate]] = None
    bank_account_ids: Optional[List[str]] = None
    validity_days: Optional[int] = None
    notes: Optional[str] = None
    include_vat: Optional[bool] = None
    # Genel İskonto
    general_discount_type: Optional[str] = None
    general_discount_value: Optional[float] = None

# Password Reset Models
class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetConfirm(BaseModel):
    token: str
    new_password: str

# Email Share Models
class EmailShareRequest(BaseModel):
    quote_id: str
    recipient_email: EmailStr
    message: Optional[str] = None

# Customer Models
class CustomerCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None
    contact_person: Optional[str] = None
    
    @field_validator('name')
    @classmethod
    def validate_name(cls, v):
        v = sanitize_input(v.strip())
        if len(v) < 1 or len(v) > 200:
            raise ValueError('Müşteri adı 1-200 karakter arası olmalıdır')
        return v
    
    @field_validator('address')
    @classmethod
    def validate_address(cls, v):
        if v:
            v = sanitize_input(v.strip())
            if len(v) > 500:
                raise ValueError('Adres çok uzun')
        return v
    
    @field_validator('phone')
    @classmethod
    def validate_phone(cls, v):
        if v:
            # Remove non-numeric chars and validate
            cleaned = re.sub(r'\D', '', v)
            if len(cleaned) < 10 or len(cleaned) > 15:
                raise ValueError('Geçersiz telefon numarası')
        return v
    
    @field_validator('tax_number')
    @classmethod
    def validate_tax_number(cls, v):
        if v:
            v = sanitize_input(v.strip())
            if len(v) > 20:
                raise ValueError('Vergi numarası çok uzun')
        return v

class CustomerResponse(BaseModel):
    id: str
    user_id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_number: Optional[str] = None
    contact_person: Optional[str] = None
    created_at: str
    updated_at: str

class QuoteItemResponse(BaseModel):
    product_id: str
    product_name: str
    unit: str
    quantity: float
    unit_price: float
    vat_rate: float
    discount_percent: float
    subtotal: float
    vat_amount: float
    total: float

class QuoteResponse(BaseModel):
    id: str
    user_id: str
    quote_number: str
    quote_name: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: str
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    customer_tax_number: Optional[str] = None
    items: List[QuoteItemResponse]
    bank_accounts: List[BankAccountResponse] = []
    subtotal: float
    total_vat: float
    total: float
    general_discount_type: Optional[str] = None
    general_discount_value: float = 0.0
    general_discount_amount: float = 0.0
    validity_date: str
    notes: Optional[str] = None
    include_vat: bool
    status: str = "draft"
    created_at: str
    updated_at: str

class SubscriptionCreate(BaseModel):
    card_holder_name: str
    card_number: str
    expire_month: str
    expire_year: str
    cvc: str

# Iyzico Checkout Form Request
class IyzicoCheckoutRequest(BaseModel):
    callback_url: str

# ============== IYZICO HELPERS ==============

import iyzipay

def get_iyzico_options():
    """Get iyzico API options"""
    options = {
        'api_key': IYZICO_API_KEY,
        'secret_key': IYZICO_SECRET_KEY,
        'base_url': IYZICO_BASE_URL
    }
    return options

async def create_iyzico_checkout_form(user: dict, callback_url: str) -> dict:
    """Initialize Iyzico checkout form for subscription"""
    options = get_iyzico_options()
    
    buyer = {
        'id': user['id'],
        'name': user.get('company_name', 'Müşteri').split()[0] if user.get('company_name') else 'Müşteri',
        'surname': user.get('company_name', 'Müşteri').split()[-1] if user.get('company_name') and len(user.get('company_name', '').split()) > 1 else 'Müşteri',
        'gsmNumber': user.get('company_phone', '+905350000000') or '+905350000000',
        'email': user['email'],
        'identityNumber': '11111111111',
        'registrationAddress': user.get('company_address', 'Türkiye') or 'Türkiye',
        'ip': '85.34.78.112',
        'city': 'Istanbul',
        'country': 'Turkey'
    }
    
    address = {
        'contactName': user.get('company_name', 'Müşteri'),
        'city': 'Istanbul',
        'country': 'Turkey',
        'address': user.get('company_address', 'Türkiye') or 'Türkiye'
    }
    
    basket_items = [
        {
            'id': 'PRO_MONTHLY',
            'name': 'TeklifMaster Aylık Abonelik',
            'category1': 'Yazılım',
            'itemType': 'VIRTUAL',
            'price': '299.00'
        }
    ]
    
    request = {
        'locale': 'tr',
        'conversationId': str(uuid.uuid4()),
        'price': '100.00',
        'paidPrice': '299.00',
        'currency': 'TRY',
        'basketId': f'BASKET-{user["id"][:8]}',
        'paymentGroup': 'SUBSCRIPTION',
        'callbackUrl': callback_url,
        'buyer': buyer,
        'shippingAddress': address,
        'billingAddress': address,
        'basketItems': basket_items
    }
    
    checkout_form_initialize = iyzipay.CheckoutFormInitialize().create(request, options)
    result = checkout_form_initialize.read().decode('utf-8')
    return json.loads(result)

async def retrieve_iyzico_checkout_result(token: str) -> dict:
    """Retrieve checkout form payment result"""
    options = get_iyzico_options()
    
    request = {
        'locale': 'tr',
        'conversationId': str(uuid.uuid4()),
        'token': token
    }
    
    checkout_form_result = iyzipay.CheckoutForm().retrieve(request, options)
    result = checkout_form_result.read().decode('utf-8')
    return json.loads(result)

# ============== AUTH HELPERS ==============

def create_token(user_id: str) -> str:
    payload = {
        "user_id": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token süresi dolmuş")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")

async def get_current_user(authorization: str = None):
    if not authorization:
        raise HTTPException(status_code=401, detail="Yetkilendirme gerekli")
    
    try:
        scheme, token = authorization.split()
        if scheme.lower() != "bearer":
            raise HTTPException(status_code=401, detail="Geçersiz yetkilendirme")
    except ValueError:
        raise HTTPException(status_code=401, detail="Geçersiz yetkilendirme formatı")
    
    payload = verify_token(token)
    user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
    return user

from fastapi import Header

async def get_auth_user(authorization: str = Header(None)):
    return await get_current_user(authorization)

# ============== AUTH ENDPOINTS ==============

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Bu email zaten kayıtlı")
    
    hashed_password = bcrypt.hashpw(user_data.password.encode(), bcrypt.gensalt()).decode()
    trial_end = datetime.now(timezone.utc) + timedelta(days=7)
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "password": hashed_password,
        "company_name": user_data.company_name,
        "company_logo": None,
        "company_address": None,
        "company_phone": None,
        "company_tax_number": None,
        "pdf_template": "classic",
        "pdf_show_images": True,
        "pdf_image_size": "medium",
        "pdf_description_length": "full",
        "trial_end_date": trial_end.isoformat(),
        "subscription_status": "trial",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    token = create_token(user_doc["id"])
    user_response = {k: v for k, v in user_doc.items() if k != "password" and k != "_id"}
    
    return {"token": token, "user": user_response}

@api_router.post("/auth/login")
async def login(credentials: UserLogin, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    
    # Check for brute force attacks
    if not check_login_attempts(credentials.email, client_ip):
        logger.warning(f"SECURITY: Too many login attempts for {credentials.email} from {client_ip}")
        block_ip(client_ip)
        raise HTTPException(
            status_code=429, 
            detail="Çok fazla başarısız giriş denemesi. 15 dakika bekleyin."
        )
    
    user = await db.users.find_one({"email": credentials.email})
    if not user:
        record_login_attempt(credentials.email, client_ip)
        raise HTTPException(status_code=401, detail="Email veya şifre hatalı")
    
    if not bcrypt.checkpw(credentials.password.encode(), user["password"].encode()):
        record_login_attempt(credentials.email, client_ip)
        logger.warning(f"SECURITY: Failed login for {credentials.email} from {client_ip}")
        raise HTTPException(status_code=401, detail="Email veya şifre hatalı")
    
    # Successful login - clear attempts
    key = f"{credentials.email}:{client_ip}"
    login_attempts.pop(key, None)
    
    logger.info(f"AUDIT: Successful login for {credentials.email} from {client_ip}")
    
    token = create_token(user["id"])
    user_response = {k: v for k, v in user.items() if k != "password" and k != "_id"}
    
    return {"token": token, "user": user_response}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_auth_user)):
    return {k: v for k, v in current_user.items() if k != "password"}

@api_router.put("/auth/profile")
async def update_profile(update_data: UserUpdate, current_user: dict = Depends(get_auth_user)):
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.users.update_one({"id": current_user["id"]}, {"$set": update_dict})
    
    updated_user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    return updated_user

@api_router.post("/auth/upload-logo")
async def upload_logo(file: UploadFile = File(...), current_user: dict = Depends(get_auth_user)):
    content = await file.read()
    encoded = base64.b64encode(content).decode()
    content_type = file.content_type or "image/png"
    data_url = f"data:{content_type};base64,{encoded}"
    
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"company_logo": data_url}})
    
    return {"logo_url": data_url}

# ============== PASSWORD RESET ==============

@api_router.post("/auth/forgot-password")
async def forgot_password(request: PasswordResetRequest):
    """Request password reset email"""
    user = await db.users.find_one({"email": request.email})
    if not user:
        # Don't reveal if email exists
        return {"message": "Şifre sıfırlama linki e-posta adresinize gönderildi"}
    
    # Generate reset token
    reset_token = secrets.token_urlsafe(32)
    expires = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Store token
    await db.password_resets.insert_one({
        "user_id": user["id"],
        "token": reset_token,
        "expires": expires.isoformat(),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://teklifmaster.com')
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    html_content = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #0F172A; padding: 20px; text-align: center;">
            <h1 style="color: #F97316; margin: 0;">TeklifMaster</h1>
        </div>
        <div style="padding: 30px; background: #f8fafc;">
            <h2 style="color: #1e293b;">Şifre Sıfırlama</h2>
            <p style="color: #475569;">Merhaba,</p>
            <p style="color: #475569;">Şifrenizi sıfırlamak için aşağıdaki butona tıklayın:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background: #F97316; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold;">Şifremi Sıfırla</a>
            </div>
            <p style="color: #64748b; font-size: 14px;">Bu link 1 saat geçerlidir.</p>
            <p style="color: #64748b; font-size: 14px;">Bu isteği siz yapmadıysanız, bu e-postayı görmezden gelebilirsiniz.</p>
        </div>
    </div>
    """
    
    # Try SMTP first (Hostinger)
    if SMTP_PASSWORD:
        try:
            send_smtp_email(request.email, "TeklifMaster - Şifre Sıfırlama", html_content)
            logging.info(f"Password reset email sent via SMTP to {request.email}")
        except Exception as e:
            logging.error(f"Failed to send password reset email via SMTP: {e}")
    # Fallback to Resend if configured
    elif RESEND_AVAILABLE and RESEND_API_KEY and RESEND_API_KEY != 're_your_api_key_here':
        try:
            params = {
                "from": SENDER_EMAIL,
                "to": [request.email],
                "subject": "TeklifMaster - Şifre Sıfırlama",
                "html": html_content
            }
            await asyncio.to_thread(resend.Emails.send, params)
        except Exception as e:
            logging.error(f"Failed to send password reset email via Resend: {e}")
    
    return {"message": "Şifre sıfırlama linki e-posta adresinize gönderildi"}

@api_router.post("/auth/reset-password")
async def reset_password(request: PasswordResetConfirm):
    """Reset password with token"""
    reset_doc = await db.password_resets.find_one({
        "token": request.token,
        "used": False
    })
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Geçersiz veya süresi dolmuş token")
    
    expires = datetime.fromisoformat(reset_doc["expires"].replace("Z", "+00:00"))
    if expires < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Token süresi dolmuş")
    
    # Update password
    hashed_password = bcrypt.hashpw(request.new_password.encode(), bcrypt.gensalt()).decode()
    await db.users.update_one(
        {"id": reset_doc["user_id"]},
        {"$set": {"password": hashed_password}}
    )
    
    # Mark token as used
    await db.password_resets.update_one(
        {"token": request.token},
        {"$set": {"used": True}}
    )
    
    return {"message": "Şifreniz başarıyla güncellendi"}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """Verify if reset token is valid"""
    reset_doc = await db.password_resets.find_one({
        "token": token,
        "used": False
    })
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Geçersiz token")
    
    expires = datetime.fromisoformat(reset_doc["expires"].replace("Z", "+00:00"))
    if expires < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Token süresi dolmuş")
    
    return {"valid": True}

# ============== PRODUCT ENDPOINTS ==============

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(current_user: dict = Depends(get_auth_user)):
    products = await db.products.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    return products

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product_data: ProductCreate, current_user: dict = Depends(get_auth_user)):
    now = datetime.now(timezone.utc).isoformat()
    product_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        **product_data.model_dump(),
        "created_at": now,
        "updated_at": now
    }
    await db.products.insert_one(product_doc)
    del product_doc["_id"]
    return product_doc

@api_router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product_data: ProductCreate, current_user: dict = Depends(get_auth_user)):
    existing = await db.products.find_one({"id": product_id, "user_id": current_user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    
    update_dict = product_data.model_dump()
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.products.update_one({"id": product_id}, {"$set": update_dict})
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    return updated

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_auth_user)):
    result = await db.products.delete_one({"id": product_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return {"message": "Ürün silindi"}

@api_router.post("/products/{product_id}/upload-image")
async def upload_product_image(product_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_auth_user)):
    existing = await db.products.find_one({"id": product_id, "user_id": current_user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    
    content = await file.read()
    encoded = base64.b64encode(content).decode()
    content_type = file.content_type or "image/png"
    data_url = f"data:{content_type};base64,{encoded}"
    
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"image_url": data_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"image_url": data_url}

# ============== EXCEL IMPORT/EXPORT ==============

@api_router.get("/products/export/excel")
async def export_products_excel(current_user: dict = Depends(get_auth_user)):
    products = await db.products.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    headers = ["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat (₺)", "KDV Oranı (%)"]
    ws.append(headers)
    
    for product in products:
        ws.append([
            product.get("sku", ""),
            product.get("name", ""),
            product.get("description", ""),
            product.get("unit", "adet"),
            product.get("unit_price", 0),
            product.get("vat_rate", 20)
        ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=urunler.xlsx"}
    )

@api_router.post("/products/import/excel")
async def import_products_excel(file: UploadFile = File(...), current_user: dict = Depends(get_auth_user)):
    content = await file.read()
    buffer = BytesIO(content)
    
    try:
        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        
        imported_count = 0
        updated_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:
                continue
            
            sku = str(row[0]) if row[0] else None
            name = str(row[1])
            description = str(row[2]) if row[2] else None
            unit = str(row[3]) if row[3] else "adet"
            unit_price = float(row[4]) if row[4] else 0
            vat_rate = float(row[5]) if row[5] else 20
            
            existing = None
            if sku:
                existing = await db.products.find_one({"sku": sku, "user_id": current_user["id"]})
            
            now = datetime.now(timezone.utc).isoformat()
            
            if existing:
                await db.products.update_one(
                    {"id": existing["id"]},
                    {"$set": {
                        "name": name,
                        "description": description,
                        "unit": unit,
                        "unit_price": unit_price,
                        "vat_rate": vat_rate,
                        "updated_at": now
                    }}
                )
                updated_count += 1
            else:
                product_doc = {
                    "id": str(uuid.uuid4()),
                    "user_id": current_user["id"],
                    "sku": sku,
                    "name": name,
                    "description": description,
                    "unit": unit,
                    "unit_price": unit_price,
                    "vat_rate": vat_rate,
                    "image_url": None,
                    "created_at": now,
                    "updated_at": now
                }
                await db.products.insert_one(product_doc)
                imported_count += 1
        
        return {"imported": imported_count, "updated": updated_count}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası okunamadı: {str(e)}")

@api_router.get("/products/template/excel")
async def download_excel_template():
    """Download empty Excel template for product import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Headers
    headers = ["SKU", "Ürün Adı", "Açıklama", "Birim", "Birim Fiyat", "KDV Oranı (%)"]
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Example rows
    example_data = [
        ["PRD-001", "Web Tasarım Hizmeti", "Kurumsal web sitesi tasarımı", "adet", 5000, 20],
        ["PRD-002", "Logo Tasarım", "Profesyonel logo tasarımı", "adet", 1500, 20],
        ["PRD-003", "SEO Danışmanlık", "Aylık SEO optimizasyon hizmeti", "ay", 2000, 20],
        ["PRD-004", "Sosyal Medya Yönetimi", "Haftalık içerik üretimi ve yönetim", "ay", 3000, 20],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [5, 6]:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Açıklamalar")
    instructions = [
        ["TeklifMaster - Ürün İçe Aktarma Şablonu"],
        [""],
        ["Kullanım Talimatları:"],
        ["1. 'Ürünler' sayfasındaki örnek verileri silin ve kendi ürünlerinizi girin"],
        ["2. SKU (Stok Kodu) alanı opsiyoneldir, ancak güncelleme için benzersiz olmalıdır"],
        ["3. Ürün Adı zorunludur"],
        ["4. Birim seçenekleri: adet, m, m2, kg, lt, saat, gun, ay, paket"],
        ["5. KDV Oranı: 0, 1, 10, 20 değerlerinden birini kullanın"],
        [""],
        ["Önemli Notlar:"],
        ["- Aynı SKU ile kayıtlı ürün varsa, bilgiler güncellenir"],
        ["- Yeni SKU veya SKU boş ise yeni ürün oluşturulur"],
        ["- İlk satır (başlık) otomatik olarak atlanır"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_idx == 3 or row_idx == 10:
            cell.font = Font(bold=True)
    
    ws2.column_dimensions['A'].width = 70
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quotemaster_urun_sablonu.xlsx"}
    )

# ============== BANK ACCOUNT ENDPOINTS ==============

@api_router.get("/bank-accounts", response_model=List[BankAccountResponse])
async def get_bank_accounts(current_user: dict = Depends(get_auth_user)):
    accounts = await db.bank_accounts.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(100)
    return accounts

@api_router.post("/bank-accounts", response_model=BankAccountResponse)
async def create_bank_account(account_data: BankAccountCreate, current_user: dict = Depends(get_auth_user)):
    account_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        **account_data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.bank_accounts.insert_one(account_doc)
    del account_doc["_id"]
    return account_doc

@api_router.delete("/bank-accounts/{account_id}")
async def delete_bank_account(account_id: str, current_user: dict = Depends(get_auth_user)):
    result = await db.bank_accounts.delete_one({"id": account_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Banka hesabı bulunamadı")
    return {"message": "Banka hesabı silindi"}

# ============== CUSTOMER ENDPOINTS ==============

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(search: str = None, current_user: dict = Depends(get_auth_user)):
    query = {"user_id": current_user["id"]}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"tax_number": {"$regex": search, "$options": "i"}}
        ]
    customers = await db.customers.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return customers

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_auth_user)):
    customer = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return customer

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer_data: CustomerCreate, current_user: dict = Depends(get_auth_user)):
    now = datetime.now(timezone.utc).isoformat()
    customer_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        **customer_data.model_dump(),
        "created_at": now,
        "updated_at": now
    }
    await db.customers.insert_one(customer_doc)
    del customer_doc["_id"]
    return customer_doc

@api_router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(customer_id: str, customer_data: CustomerCreate, current_user: dict = Depends(get_auth_user)):
    existing = await db.customers.find_one({"id": customer_id, "user_id": current_user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    update_dict = customer_data.model_dump()
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.customers.update_one({"id": customer_id}, {"$set": update_dict})
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return updated

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_auth_user)):
    result = await db.customers.delete_one({"id": customer_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return {"message": "Müşteri silindi"}

# ============== QUOTE ENDPOINTS ==============

async def get_next_quote_number(user_id: str) -> str:
    last_quote = await db.quotes.find_one(
        {"user_id": user_id},
        sort=[("created_at", -1)]
    )
    
    year = datetime.now().year
    if last_quote and last_quote.get("quote_number", "").startswith(f"TKL-{year}"):
        try:
            last_num = int(last_quote["quote_number"].split("-")[-1])
            return f"TKL-{year}-{str(last_num + 1).zfill(4)}"
        except:
            pass
    
    return f"TKL-{year}-0001"

@api_router.get("/quotes", response_model=List[QuoteResponse])
async def get_quotes(current_user: dict = Depends(get_auth_user)):
    quotes = await db.quotes.find({"user_id": current_user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return quotes

@api_router.get("/quotes/{quote_id}", response_model=QuoteResponse)
async def get_quote(quote_id: str, current_user: dict = Depends(get_auth_user)):
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user["id"]}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    return quote

@api_router.post("/quotes", response_model=QuoteResponse)
async def create_quote(quote_data: QuoteCreate, current_user: dict = Depends(get_auth_user)):
    now = datetime.now(timezone.utc)
    validity_date = now + timedelta(days=quote_data.validity_days)
    
    # Calculate items with totals
    items = []
    subtotal = 0
    total_vat = 0
    
    for item in quote_data.items:
        item_subtotal = item.quantity * item.unit_price
        discount_amount = item_subtotal * (item.discount_percent / 100)
        item_subtotal_after_discount = item_subtotal - discount_amount
        vat_amount = item_subtotal_after_discount * (item.vat_rate / 100) if quote_data.include_vat else 0
        item_total = item_subtotal_after_discount + vat_amount
        
        items.append({
            "product_id": item.product_id,
            "product_name": item.product_name,
            "unit": item.unit,
            "quantity": item.quantity,
            "unit_price": item.unit_price,
            "vat_rate": item.vat_rate,
            "discount_percent": item.discount_percent,
            "subtotal": round(item_subtotal_after_discount, 2),
            "vat_amount": round(vat_amount, 2),
            "total": round(item_total, 2)
        })
        
        subtotal += item_subtotal_after_discount
        total_vat += vat_amount
    
    # Get bank accounts
    bank_accounts = []
    if quote_data.bank_account_ids:
        accounts = await db.bank_accounts.find(
            {"id": {"$in": quote_data.bank_account_ids}, "user_id": current_user["id"]},
            {"_id": 0}
        ).to_list(10)
        bank_accounts = accounts
    
    quote_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "quote_number": await get_next_quote_number(current_user["id"]),
        "quote_name": quote_data.quote_name,
        "customer_id": quote_data.customer_id,
        "customer_name": quote_data.customer_name,
        "customer_email": quote_data.customer_email,
        "customer_phone": quote_data.customer_phone,
        "customer_address": quote_data.customer_address,
        "customer_tax_number": quote_data.customer_tax_number,
        "items": items,
        "bank_accounts": bank_accounts,
        "subtotal": round(subtotal, 2),
        "total_vat": round(total_vat, 2),
        "general_discount_type": quote_data.general_discount_type,
        "general_discount_value": quote_data.general_discount_value,
        "general_discount_amount": 0,
        "total": round(subtotal + total_vat, 2),
        "validity_date": validity_date.isoformat(),
        "notes": quote_data.notes,
        "include_vat": quote_data.include_vat,
        "status": "draft",
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    # Apply general discount
    if quote_data.general_discount_type and quote_data.general_discount_value > 0:
        base_total = subtotal + total_vat
        if quote_data.general_discount_type == "percent":
            general_discount_amount = base_total * (quote_data.general_discount_value / 100)
        else:  # amount
            general_discount_amount = quote_data.general_discount_value
        quote_doc["general_discount_amount"] = round(general_discount_amount, 2)
        quote_doc["total"] = round(base_total - general_discount_amount, 2)
    
    await db.quotes.insert_one(quote_doc)
    del quote_doc["_id"]
    return quote_doc

@api_router.put("/quotes/{quote_id}/status")
async def update_quote_status(quote_id: str, status: str, current_user: dict = Depends(get_auth_user)):
    valid_statuses = ["draft", "sent", "accepted", "rejected"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Geçersiz durum")
    
    result = await db.quotes.update_one(
        {"id": quote_id, "user_id": current_user["id"]},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    
    return {"message": "Durum güncellendi"}

@api_router.put("/quotes/{quote_id}", response_model=QuoteResponse)
async def update_quote(quote_id: str, quote_data: QuoteUpdate, current_user: dict = Depends(get_auth_user)):
    existing = await db.quotes.find_one({"id": quote_id, "user_id": current_user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    
    update_dict = {}
    
    # Update basic fields
    for field in ["quote_name", "customer_id", "customer_name", "customer_email", "customer_phone", "customer_address", "customer_tax_number", "notes", "include_vat"]:
        value = getattr(quote_data, field, None)
        if value is not None:
            update_dict[field] = value
    
    # Update validity date if days provided
    if quote_data.validity_days is not None:
        created_at = datetime.fromisoformat(existing["created_at"].replace("Z", "+00:00"))
        validity_date = created_at + timedelta(days=quote_data.validity_days)
        update_dict["validity_date"] = validity_date.isoformat()
    
    # Update bank accounts
    if quote_data.bank_account_ids is not None:
        bank_accounts = []
        if quote_data.bank_account_ids:
            accounts = await db.bank_accounts.find(
                {"id": {"$in": quote_data.bank_account_ids}, "user_id": current_user["id"]},
                {"_id": 0}
            ).to_list(10)
            bank_accounts = accounts
        update_dict["bank_accounts"] = bank_accounts
    
    # Update items if provided
    if quote_data.items is not None:
        include_vat = quote_data.include_vat if quote_data.include_vat is not None else existing.get("include_vat", True)
        items = []
        subtotal = 0
        total_vat = 0
        
        for item in quote_data.items:
            item_subtotal = item.quantity * item.unit_price
            discount_amount = item_subtotal * (item.discount_percent / 100)
            item_subtotal_after_discount = item_subtotal - discount_amount
            vat_amount = item_subtotal_after_discount * (item.vat_rate / 100) if include_vat else 0
            item_total = item_subtotal_after_discount + vat_amount
            
            items.append({
                "product_id": item.product_id,
                "product_name": item.product_name,
                "unit": item.unit,
                "quantity": item.quantity,
                "unit_price": item.unit_price,
                "vat_rate": item.vat_rate,
                "discount_percent": item.discount_percent,
                "subtotal": round(item_subtotal_after_discount, 2),
                "vat_amount": round(vat_amount, 2),
                "total": round(item_total, 2)
            })
            
            subtotal += item_subtotal_after_discount
            total_vat += vat_amount
        
        update_dict["items"] = items
        update_dict["subtotal"] = round(subtotal, 2)
        update_dict["total_vat"] = round(total_vat, 2)
        update_dict["total"] = round(subtotal + total_vat, 2)
    
    # Update general discount
    if quote_data.general_discount_type is not None:
        update_dict["general_discount_type"] = quote_data.general_discount_type
    if quote_data.general_discount_value is not None:
        update_dict["general_discount_value"] = quote_data.general_discount_value
    
    # Recalculate total with general discount
    final_subtotal = update_dict.get("subtotal", existing.get("subtotal", 0))
    final_vat = update_dict.get("total_vat", existing.get("total_vat", 0))
    discount_type = update_dict.get("general_discount_type", existing.get("general_discount_type"))
    discount_value = update_dict.get("general_discount_value", existing.get("general_discount_value", 0))
    
    if discount_type and discount_value and discount_value > 0:
        base_total = final_subtotal + final_vat
        if discount_type == "percent":
            general_discount_amount = base_total * (discount_value / 100)
        else:
            general_discount_amount = discount_value
        update_dict["general_discount_amount"] = round(general_discount_amount, 2)
        update_dict["total"] = round(base_total - general_discount_amount, 2)
    
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.quotes.update_one({"id": quote_id}, {"$set": update_dict})
    updated = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    return updated

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str, current_user: dict = Depends(get_auth_user)):
    result = await db.quotes.delete_one({"id": quote_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    return {"message": "Teklif silindi"}

# ============== PDF GENERATION ==============

# ============== PDF TEMPLATES ==============

PDF_TEMPLATES = {
    "classic": {
        "name": "Klasik",
        "primary": "#1E3A5F",
        "accent": "#F97316",
        "bg": "#F8FAFC",
        "border": "#E2E8F0",
        "text": "#374151",
        "muted": "#64748B"
    },
    "modern": {
        "name": "Modern",
        "primary": "#0F172A",
        "accent": "#F97316",
        "bg": "#FFF7ED",
        "border": "#FED7AA",
        "text": "#374151",
        "muted": "#6B7280"
    },
    "professional": {
        "name": "Profesyonel",
        "primary": "#1F2937",
        "accent": "#1F2937",
        "bg": "#F9FAFB",
        "border": "#D1D5DB",
        "text": "#111827",
        "muted": "#6B7280"
    },
    "elegant": {
        "name": "Zarif",
        "primary": "#78350F",
        "accent": "#B45309",
        "bg": "#FFFBEB",
        "border": "#FDE68A",
        "text": "#451A03",
        "muted": "#92400E"
    },
    "ocean": {
        "name": "Okyanus",
        "primary": "#0369A1",
        "accent": "#0891B2",
        "bg": "#F0F9FF",
        "border": "#BAE6FD",
        "text": "#0C4A6E",
        "muted": "#0369A1"
    }
}

def format_currency_pdf(amount):
    """Format currency for PDF"""
    return f"{amount:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")

def generate_pdf_with_template(quote, user, template_id="classic", pdf_settings=None):
    """Generate PDF with specified template and settings"""
    template = PDF_TEMPLATES.get(template_id, PDF_TEMPLATES["classic"])
    
    # Register Turkish-compatible fonts - try multiple font paths
    font_paths = [
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf', '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
        ('/usr/share/fonts/TTF/DejaVuSans.ttf', '/usr/share/fonts/TTF/DejaVuSans-Bold.ttf'),
    ]
    
    default_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'
    
    for regular_path, bold_path in font_paths:
        try:
            if os.path.exists(regular_path) and os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont('TurkishFont', regular_path))
                pdfmetrics.registerFont(TTFont('TurkishFont-Bold', bold_path))
                default_font = 'TurkishFont'
                bold_font = 'TurkishFont-Bold'
                logging.info(f"Loaded Turkish font from {regular_path}")
                break
        except Exception as e:
            logging.warning(f"Could not load font from {regular_path}: {e}")
            continue
    
    # PDF Settings defaults
    if pdf_settings is None:
        pdf_settings = {}
    show_images = pdf_settings.get('pdf_show_images', user.get('pdf_show_images', True))
    image_size = pdf_settings.get('pdf_image_size', user.get('pdf_image_size', 'medium'))
    description_length = pdf_settings.get('pdf_description_length', user.get('pdf_description_length', 'full'))
    
    # Image size mapping
    image_sizes = {'small': 1*cm, 'medium': 1.5*cm, 'large': 2*cm}
    img_size = image_sizes.get(image_size, 1.5*cm)
    
    # Format dates
    created_date = datetime.fromisoformat(quote["created_at"].replace("Z", "+00:00")).strftime("%d.%m.%Y")
    validity_date = datetime.fromisoformat(quote["validity_date"].replace("Z", "+00:00")).strftime("%d.%m.%Y")
    
    # Create PDF buffer
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    # Styles based on template with Turkish font
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Title2', fontSize=18, fontName=bold_font, textColor=colors.HexColor(template['primary']), spaceAfter=10))
    styles.add(ParagraphStyle(name='Subtitle', fontSize=10, fontName=default_font, textColor=colors.HexColor(template['muted'])))
    styles.add(ParagraphStyle(name='QuoteNumber', fontSize=14, fontName=bold_font, textColor=colors.HexColor(template['accent']), alignment=2))
    styles.add(ParagraphStyle(name='SectionTitle', fontSize=11, fontName=bold_font, textColor=colors.HexColor(template['primary']), spaceBefore=20, spaceAfter=10))
    styles.add(ParagraphStyle(name='Normal2', fontSize=9, fontName=default_font, textColor=colors.HexColor(template['text'])))
    styles.add(ParagraphStyle(name='Footer', fontSize=8, fontName=default_font, textColor=colors.HexColor(template['muted']), alignment=1))
    styles.add(ParagraphStyle(name='FooterLink', fontSize=8, fontName=default_font, textColor=colors.HexColor('#F97316'), alignment=1))
    
    elements = []
    
    # Header Section with Logo
    company_logo = user.get('company_logo')
    header_left_content = []
    
    if company_logo and company_logo.startswith('data:'):
        try:
            # Decode base64 logo
            header_data = company_logo.split(',')[1]
            logo_data = base64.b64decode(header_data)
            logo_buffer = BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=3*cm, height=3*cm, kind='proportional')
            header_left_content.append(logo_img)
            header_left_content.append(Spacer(1, 5))
        except Exception as e:
            logging.error(f"Failed to load logo: {e}")
    
    header_left_content.append(Paragraph(f"<b>{user.get('company_name', 'Şirket')}</b>", styles['Title2']))
    
    header_data = [
        [
            header_left_content,
            Paragraph(quote['quote_number'], styles['QuoteNumber'])
        ],
        [
            Paragraph(f"{user.get('company_address', '') or ''}<br/>{user.get('company_phone', '') or ''}<br/>{'VKN: ' + user.get('company_tax_number') if user.get('company_tax_number') else ''}", styles['Subtitle']),
            Paragraph(f"<b>Tarih:</b> {created_date}<br/><b>Geçerlilik:</b> {validity_date}", styles['Subtitle'])
        ]
    ]
    
    header_table = Table(header_data, colWidths=[10*cm, 6*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 1), (-1, 1), 2, colors.HexColor(template['accent'])),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Customer Section
    elements.append(Paragraph("MÜŞTERİ BİLGİLERİ", styles['SectionTitle']))
    customer_info = f"<b>{quote['customer_name']}</b><br/>"
    if quote.get('customer_tax_number'):
        customer_info += f"VKN: {quote['customer_tax_number']}<br/>"
    if quote.get('customer_address'):
        customer_info += f"{quote['customer_address']}<br/>"
    if quote.get('customer_phone'):
        customer_info += f"Tel: {quote['customer_phone']}<br/>"
    if quote.get('customer_email'):
        customer_info += f"Email: {quote['customer_email']}"
    
    customer_table = Table([[Paragraph(customer_info, styles['Normal2'])]], colWidths=[16*cm])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(template['bg'])),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(template['border'])),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 20))
    
    # Items Section
    elements.append(Paragraph("TEKLIF DETAYLARI", styles['SectionTitle']))
    
    items_data = [['#', 'Urun/Hizmet', 'Miktar', 'Birim Fiyat', 'Iskonto', 'KDV', 'Toplam']]
    
    for idx, item in enumerate(quote["items"], 1):
        discount_str = f"%{item['discount_percent']}" if item["discount_percent"] > 0 else "-"
        items_data.append([
            str(idx),
            item['product_name'],
            f"{item['quantity']} {item['unit']}",
            format_currency_pdf(item['unit_price']),
            discount_str,
            f"%{int(item['vat_rate'])}",
            format_currency_pdf(item['total'])
        ])
    
    items_table = Table(items_data, colWidths=[1*cm, 5*cm, 2*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2.5*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(template['primary'])),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(template['border'])),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(template['bg'])]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 10))
    
    # Totals Section with General Discount
    totals_data = [
        ['Ara Toplam:', format_currency_pdf(quote['subtotal'])],
        ['KDV Toplami:', format_currency_pdf(quote['total_vat'])],
    ]
    
    # Add general discount if exists
    if quote.get('general_discount_type') and quote.get('general_discount_amount', 0) > 0:
        discount_label = f"Genel İskonto ({quote['general_discount_value']}{'%' if quote['general_discount_type'] == 'percent' else ' TL'}):"
        totals_data.append([discount_label, f"-{format_currency_pdf(quote['general_discount_amount'])}"])
    
    totals_data.append(['GENEL TOPLAM:', format_currency_pdf(quote['total'])])
    
    totals_table = Table(totals_data, colWidths=[10*cm, 6*cm])
    
    # Style based on number of rows
    total_row = len(totals_data) - 1
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, total_row - 1), 'Helvetica'),
        ('FONTNAME', (0, total_row), (-1, total_row), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, total_row - 1), 9),
        ('FONTSIZE', (0, total_row), (-1, total_row), 12),
        ('TEXTCOLOR', (1, total_row), (1, total_row), colors.HexColor(template['accent'])),
        ('LINEABOVE', (0, total_row), (-1, total_row), 2, colors.HexColor(template['primary'])),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(totals_table)
    
    # Bank Accounts Section
    if quote.get("bank_accounts") and len(quote["bank_accounts"]) > 0:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("ÖDEME BİLGİLERİ", styles['SectionTitle']))
        
        for account in quote["bank_accounts"]:
            currency_symbol = "₺" if account.get('currency', 'TRY') == 'TRY' else account.get('currency', '')
            bank_info = f"<b>{account['bank_name']}</b> ({currency_symbol})<br/>IBAN: {account['iban']}"
            if account.get('account_holder'):
                bank_info += f"<br/>Hesap Sahibi: {account['account_holder']}"
            
            bank_table = Table([[Paragraph(bank_info, styles['Normal2'])]], colWidths=[16*cm])
            bank_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(template['bg'])),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(template['border'])),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(bank_table)
            elements.append(Spacer(1, 5))
    
    # Notes Section
    if quote.get('notes'):
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("NOTLAR", styles['SectionTitle']))
        notes_table = Table([[Paragraph(quote['notes'], styles['Normal2'])]], colWidths=[16*cm])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFBEB')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#FDE68A')),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)
    
    # Footer with link
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Bu teklif {validity_date} tarihine kadar geçerlidir.", styles['Footer']))
    elements.append(Paragraph('<a href="https://teklifmaster.com" color="#F97316">teklifmaster.com</a> ile oluşturuldu.', styles['FooterLink']))
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return pdf_buffer

@api_router.get("/pdf-templates")
async def get_pdf_templates(current_user: dict = Depends(get_auth_user)):
    """Get available PDF templates"""
    templates = []
    for key, value in PDF_TEMPLATES.items():
        templates.append({
            "id": key,
            "name": value["name"],
            "primary": value["primary"],
            "accent": value["accent"]
        })
    return templates

@api_router.get("/quotes/{quote_id}/pdf")
async def generate_quote_pdf(quote_id: str, current_user: dict = Depends(get_auth_user)):
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user["id"]}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    
    # Get user's selected template
    template_id = user.get("pdf_template", "classic")
    
    # Generate PDF with template
    pdf_buffer = generate_pdf_with_template(quote, user, template_id)
    
    filename = f"Teklif_{quote['quote_number']}_{quote['customer_name'].replace(' ', '_')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============== EMAIL SHARING ==============

@api_router.post("/quotes/{quote_id}/share/email")
async def share_quote_email(quote_id: str, request: EmailShareRequest, current_user: dict = Depends(get_auth_user)):
    """Share quote via email with PDF attachment"""
    if not RESEND_API_KEY or RESEND_API_KEY == 're_your_api_key_here':
        raise HTTPException(status_code=400, detail="Email servisi yapılandırılmamış. Lütfen RESEND_API_KEY ekleyin.")
    
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user["id"]}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    
    # Generate PDF
    template_id = user.get("pdf_template", "classic")
    pdf_buffer = generate_pdf_with_template(quote, user, template_id)
    pdf_content = pdf_buffer.getvalue()
    pdf_base64 = base64.b64encode(pdf_content).decode()
    
    # Build email HTML
    custom_message = request.message or ""
    html_content = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #0F172A; padding: 20px; text-align: center;">
            <h1 style="color: #F97316; margin: 0;">{user.get('company_name', 'TeklifMaster')}</h1>
        </div>
        <div style="padding: 30px; background: #f8fafc;">
            <h2 style="color: #1e293b;">Fiyat Teklifi</h2>
            <p style="color: #475569;">Sayın {quote['customer_name']},</p>
            <p style="color: #475569;">{quote['quote_number']} numaralı teklifimizi ekte bulabilirsiniz.</p>
            {f'<p style="color: #475569;">{custom_message}</p>' if custom_message else ''}
            <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 8px 0; color: #64748b;">Teklif No:</td>
                        <td style="padding: 8px 0; font-weight: bold; text-align: right;">{quote['quote_number']}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #64748b;">Toplam:</td>
                        <td style="padding: 8px 0; font-weight: bold; color: #F97316; text-align: right;">{format_currency_pdf(quote['total'])}</td>
                    </tr>
                </table>
            </div>
            <p style="color: #475569;">Sorularınız için bize ulaşabilirsiniz.</p>
            <p style="color: #475569;">İyi günler dileriz.</p>
            <p style="color: #64748b; margin-top: 30px; font-size: 12px;">
                {user.get('company_name', '')}<br/>
                {user.get('company_phone', '') or ''}<br/>
                {user.get('company_address', '') or ''}
            </p>
        </div>
    </div>
    """
    
    try:
        filename = f"Teklif_{quote['quote_number']}.pdf"
        params = {
            "from": SENDER_EMAIL,
            "to": [request.recipient_email],
            "subject": f"{quote['quote_number']} - {user.get('company_name', 'TeklifMaster')} Fiyat Teklifi",
            "html": html_content,
            "attachments": [
                {
                    "filename": filename,
                    "content": pdf_base64,
                }
            ]
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        
        # Update quote status to sent if it was draft
        if quote.get('status') == 'draft':
            await db.quotes.update_one(
                {"id": quote_id},
                {"$set": {"status": "sent", "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        
        return {
            "status": "success",
            "message": f"Teklif {request.recipient_email} adresine gönderildi",
            "email_id": result.get("id") if isinstance(result, dict) else str(result)
        }
    except Exception as e:
        logging.error(f"Failed to send quote email: {e}")
        raise HTTPException(status_code=500, detail=f"Email gönderilemedi: {str(e)}")

# ============== DASHBOARD STATS ==============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_auth_user)):
    total_products = await db.products.count_documents({"user_id": current_user["id"]})
    total_quotes = await db.quotes.count_documents({"user_id": current_user["id"]})
    
    # This month's quotes
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    monthly_quotes = await db.quotes.find({
        "user_id": current_user["id"],
        "created_at": {"$gte": month_start.isoformat()}
    }, {"_id": 0, "total": 1}).to_list(1000)
    
    monthly_total = sum(q.get("total", 0) for q in monthly_quotes)
    
    # Status counts
    status_counts = {
        "draft": await db.quotes.count_documents({"user_id": current_user["id"], "status": "draft"}),
        "sent": await db.quotes.count_documents({"user_id": current_user["id"], "status": "sent"}),
        "accepted": await db.quotes.count_documents({"user_id": current_user["id"], "status": "accepted"}),
        "rejected": await db.quotes.count_documents({"user_id": current_user["id"], "status": "rejected"})
    }
    
    # Recent quotes
    recent_quotes = await db.quotes.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "total_products": total_products,
        "total_quotes": total_quotes,
        "monthly_total": round(monthly_total, 2),
        "monthly_quote_count": len(monthly_quotes),
        "status_counts": status_counts,
        "recent_quotes": recent_quotes
    }

# ============== SUBSCRIPTION (MOCK IYZICO) ==============

@api_router.get("/subscription/status")
async def get_subscription_status(current_user: dict = Depends(get_auth_user)):
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    
    trial_end = datetime.fromisoformat(user["trial_end_date"].replace("Z", "+00:00")) if user.get("trial_end_date") else None
    is_trial_active = trial_end and trial_end > datetime.now(timezone.utc) if trial_end else False
    
    subscription = await db.subscriptions.find_one({"user_id": current_user["id"]}, {"_id": 0})
    
    return {
        "status": user.get("subscription_status", "trial"),
        "trial_end_date": user.get("trial_end_date"),
        "is_trial_active": is_trial_active,
        "subscription": subscription,
        "plan": {
            "name": "Pro",
            "price": 100,
            "currency": "TRY",
            "period": "monthly"
        }
    }

@api_router.post("/subscription/subscribe")
async def create_subscription(card_data: SubscriptionCreate, current_user: dict = Depends(get_auth_user)):
    """Initialize 3D Secure payment for subscription (100 TL monthly)"""
    options = get_iyzico_options()
    
    if not IYZICO_API_KEY or not IYZICO_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Ödeme sistemi yapılandırılmamış")
    
    # Use the frontend URL for callback
    frontend_url = os.environ.get('FRONTEND_URL', 'https://teklifmaster.com')
    callback_url = f"{frontend_url}/subscription/callback"
    
    # Prepare 3D Secure payment request
    payment_card = {
        'cardHolderName': card_data.card_holder_name,
        'cardNumber': card_data.card_number.replace(' ', ''),
        'expireMonth': card_data.expire_month,
        'expireYear': card_data.expire_year,
        'cvc': card_data.cvc,
        'registerCard': '0'
    }
    
    buyer = {
        'id': current_user['id'],
        'name': current_user.get('company_name', 'Müşteri').split()[0] if current_user.get('company_name') else 'Müşteri',
        'surname': current_user.get('company_name', 'Müşteri').split()[-1] if current_user.get('company_name') and len(current_user.get('company_name', '').split()) > 1 else 'Müşteri',
        'gsmNumber': current_user.get('company_phone', '+905350000000') or '+905350000000',
        'email': current_user['email'],
        'identityNumber': '11111111111',
        'registrationAddress': current_user.get('company_address', 'Türkiye') or 'Türkiye',
        'ip': '85.34.78.112',
        'city': 'Istanbul',
        'country': 'Turkey'
    }
    
    address = {
        'contactName': current_user.get('company_name', 'Müşteri'),
        'city': 'Istanbul',
        'country': 'Turkey',
        'address': current_user.get('company_address', 'Türkiye') or 'Türkiye'
    }
    
    basket_items = [
        {
            'id': 'PRO_MONTHLY',
            'name': 'TeklifMaster Aylık Abonelik',
            'category1': 'Yazılım',
            'itemType': 'VIRTUAL',
            'price': '299.00'
        }
    ]
    
    conversation_id = str(uuid.uuid4())
    
    request = {
        'locale': 'tr',
        'conversationId': conversation_id,
        'price': '100.00',
        'paidPrice': '299.00',
        'currency': 'TRY',
        'installment': '1',
        'basketId': f'SUB-{current_user["id"][:8]}',
        'paymentChannel': 'WEB',
        'paymentGroup': 'SUBSCRIPTION',
        'paymentCard': payment_card,
        'buyer': buyer,
        'shippingAddress': address,
        'billingAddress': address,
        'basketItems': basket_items,
        'callbackUrl': callback_url
    }
    
    try:
        # Initialize 3D Secure payment
        threeds_initialize = iyzipay.ThreedsInitialize().create(request, options)
        result = json.loads(threeds_initialize.read().decode('utf-8'))
        
        logger.info(f"Iyzico 3DS init result: {result.get('status')} - {result.get('errorMessage', 'OK')}")
        
        if result.get('status') != 'success':
            error_msg = result.get('errorMessage', 'Ödeme başlatılamadı')
            logger.error(f"Iyzico 3DS init failed: {error_msg}")
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Store pending payment session
        pending_payment = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "conversation_id": conversation_id,
            "card_last_four": card_data.card_number[-4:],
            "card_holder": card_data.card_holder_name,
            "status": "pending_3ds",
            "amount": 100,
            "currency": "TRY",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.pending_payments.insert_one(pending_payment)
        
        # Return the 3D Secure HTML content
        return {
            "status": "3ds_required",
            "threeds_html_content": result.get('threeDSHtmlContent'),
            "conversation_id": conversation_id,
            "message": "3D Secure doğrulama gerekiyor"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Iyzico 3DS init error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ödeme başlatılamadı: {str(e)}")

@api_router.post("/subscription/callback")
async def subscription_callback(request: Request):
    """Handle 3D Secure callback from Iyzico"""
    try:
        # Get form data from POST request
        form_data = await request.form()
        
        # The callback includes payment result
        options = get_iyzico_options()
        
        # Retrieve the 3DS payment result
        retrieve_request = {
            'locale': 'tr',
            'conversationId': str(uuid.uuid4()),
            'paymentId': form_data.get('paymentId')
        }
        
        payment_result = iyzipay.ThreedsPayment().create(retrieve_request, options)
        result = json.loads(payment_result.read().decode('utf-8'))
        
        logger.info(f"Iyzico 3DS callback result: {result.get('status')}")
        
        if result.get('status') == 'success':
            # Find the pending payment
            conversation_id = result.get('conversationId')
            pending = await db.pending_payments.find_one({"conversation_id": conversation_id})
            
            if pending:
                user_id = pending["user_id"]
                now = datetime.now(timezone.utc)
                next_payment = now + timedelta(days=30)
                
                subscription_doc = {
                    "id": str(uuid.uuid4()),
                    "user_id": user_id,
                    "iyzico_payment_id": result.get('paymentId'),
                    "card_last_four": pending.get("card_last_four", "****"),
                    "card_holder": pending.get("card_holder"),
                    "card_type": result.get('cardType', 'UNKNOWN'),
                    "card_association": result.get('cardAssociation', 'UNKNOWN'),
                    "status": "active",
                    "plan_name": "Pro",
                    "amount": 100,
                    "currency": "TRY",
                    "period": "monthly",
                    "paid_price": float(result.get('paidPrice', 100)),
                    "next_payment_date": next_payment.isoformat(),
                    "created_at": now.isoformat()
                }
                
                await db.subscriptions.insert_one(subscription_doc)
                await db.users.update_one(
                    {"id": user_id},
                    {"$set": {"subscription_status": "active"}}
                )
                await db.pending_payments.delete_one({"conversation_id": conversation_id})
                
                # Redirect to success page
                frontend_url = os.environ.get('FRONTEND_URL', 'https://teklifmaster.com')
                return HTMLResponse(content=f"""
                    <html>
                    <head>
                        <meta http-equiv="refresh" content="0;url={frontend_url}/subscription?status=success">
                    </head>
                    <body>
                        <p>Ödeme başarılı! Yönlendiriliyorsunuz...</p>
                    </body>
                    </html>
                """)
        
        # Payment failed
        frontend_url = os.environ.get('FRONTEND_URL', 'https://teklifmaster.com')
        error_msg = result.get('errorMessage', 'Ödeme başarısız')
        return HTMLResponse(content=f"""
            <html>
            <head>
                <meta http-equiv="refresh" content="0;url={frontend_url}/subscription?status=failed&error={error_msg}">
            </head>
            <body>
                <p>Ödeme başarısız! Yönlendiriliyorsunuz...</p>
            </body>
            </html>
        """)
        
    except Exception as e:
        logger.error(f"Subscription callback error: {str(e)}")
        frontend_url = os.environ.get('FRONTEND_URL', 'https://teklifmaster.com')
        return HTMLResponse(content=f"""
            <html>
            <head>
                <meta http-equiv="refresh" content="0;url={frontend_url}/subscription?status=error">
            </head>
            <body>
                <p>Bir hata oluştu! Yönlendiriliyorsunuz...</p>
            </body>
            </html>
        """)

@api_router.post("/subscription/checkout/initialize")
async def initialize_checkout(request: IyzicoCheckoutRequest, current_user: dict = Depends(get_auth_user)):
    """Initialize Iyzico checkout form (alternative to direct card payment)"""
    if not IYZICO_API_KEY or not IYZICO_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Ödeme sistemi yapılandırılmamış")
    
    try:
        result = await create_iyzico_checkout_form(current_user, request.callback_url)
        
        if result.get('status') != 'success':
            raise HTTPException(status_code=400, detail=result.get('errorMessage', 'Checkout başlatılamadı'))
        
        # Store checkout session
        checkout_session = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "token": result.get('token'),
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.checkout_sessions.insert_one(checkout_session)
        
        return {
            "status": "success",
            "token": result.get('token'),
            "checkout_form_content": result.get('checkoutFormContent'),
            "payment_page_url": result.get('paymentPageUrl')
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Checkout initialization error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Checkout başlatılamadı: {str(e)}")

@api_router.get("/subscription/checkout/callback")
async def checkout_callback(token: str):
    """Handle Iyzico checkout callback"""
    try:
        result = await retrieve_iyzico_checkout_result(token)
        
        if result.get('status') != 'success' or result.get('paymentStatus') != 'SUCCESS':
            return {"status": "failed", "message": result.get('errorMessage', 'Ödeme başarısız')}
        
        # Find checkout session
        session = await db.checkout_sessions.find_one({"token": token})
        if not session:
            raise HTTPException(status_code=404, detail="Checkout oturumu bulunamadı")
        
        user = await db.users.find_one({"id": session["user_id"]})
        if not user:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
        now = datetime.now(timezone.utc)
        next_payment = now + timedelta(days=30)
        
        subscription_doc = {
            "id": str(uuid.uuid4()),
            "user_id": session["user_id"],
            "iyzico_payment_id": result.get('paymentId'),
            "card_last_four": result.get('lastFourDigits', '****'),
            "status": "active",
            "plan_name": "Pro",
            "amount": 100,
            "currency": "TRY",
            "period": "monthly",
            "paid_price": float(result.get('paidPrice', 100)),
            "next_payment_date": next_payment.isoformat(),
            "created_at": now.isoformat()
        }
        
        await db.subscriptions.insert_one(subscription_doc)
        await db.users.update_one(
            {"id": session["user_id"]},
            {"$set": {"subscription_status": "active"}}
        )
        await db.checkout_sessions.update_one(
            {"token": token},
            {"$set": {"status": "completed"}}
        )
        
        return {"status": "success", "message": "Abonelik başarıyla oluşturuldu"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Checkout callback error: {str(e)}")
        return {"status": "failed", "message": str(e)}

@api_router.post("/subscription/cancel")
async def cancel_subscription(current_user: dict = Depends(get_auth_user)):
    result = await db.subscriptions.update_one(
        {"user_id": current_user["id"], "status": "active"},
        {"$set": {"status": "cancelled"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Aktif abonelik bulunamadı")
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"subscription_status": "cancelled"}}
    )
    
    return {"message": "Abonelik iptal edildi"}

# ============== REPORTS ==============

@api_router.get("/reports")
async def get_reports(start_date: str = None, end_date: str = None, current_user: dict = Depends(get_auth_user)):
    query = {"user_id": current_user["id"]}
    
    # Parse date range
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
            query["created_at"] = {"$gte": start.isoformat(), "$lte": end.isoformat()}
        except ValueError:
            raise HTTPException(status_code=400, detail="Geçersiz tarih formatı. YYYY-MM-DD kullanın.")
    
    # Get all quotes in date range
    quotes = await db.quotes.find(query, {"_id": 0}).to_list(10000)
    
    # Calculate statistics
    total_quotes = len(quotes)
    
    status_counts = {
        "draft": 0,
        "sent": 0,
        "accepted": 0,
        "rejected": 0
    }
    
    total_value = 0
    accepted_value = 0
    pending_value = 0
    rejected_value = 0
    
    for quote in quotes:
        status = quote.get("status", "draft")
        status_counts[status] = status_counts.get(status, 0) + 1
        total_value += quote.get("total", 0)
        
        if status == "accepted":
            accepted_value += quote.get("total", 0)
        elif status in ["draft", "sent"]:
            pending_value += quote.get("total", 0)
        elif status == "rejected":
            rejected_value += quote.get("total", 0)
    
    # Conversion rate
    conversion_rate = (status_counts["accepted"] / total_quotes * 100) if total_quotes > 0 else 0
    
    # Top customers by accepted quote value
    customer_values = {}
    for quote in quotes:
        if quote.get("status") == "accepted":
            customer = quote.get("customer_name", "Bilinmeyen")
            customer_values[customer] = customer_values.get(customer, 0) + quote.get("total", 0)
    
    top_customers = sorted(
        [{"name": k, "total": v} for k, v in customer_values.items()],
        key=lambda x: x["total"],
        reverse=True
    )[:10]
    
    # Monthly breakdown
    monthly_data = {}
    for quote in quotes:
        created = quote.get("created_at", "")[:7]  # YYYY-MM
        if created not in monthly_data:
            monthly_data[created] = {"quotes": 0, "accepted": 0, "total": 0, "accepted_value": 0}
        monthly_data[created]["quotes"] += 1
        monthly_data[created]["total"] += quote.get("total", 0)
        if quote.get("status") == "accepted":
            monthly_data[created]["accepted"] += 1
            monthly_data[created]["accepted_value"] += quote.get("total", 0)
    
    monthly_breakdown = [
        {"month": k, **v} for k, v in sorted(monthly_data.items())
    ]
    
    return {
        "date_range": {
            "start": start_date,
            "end": end_date
        },
        "summary": {
            "total_quotes": total_quotes,
            "status_counts": status_counts,
            "total_value": round(total_value, 2),
            "accepted_value": round(accepted_value, 2),
            "pending_value": round(pending_value, 2),
            "rejected_value": round(rejected_value, 2),
            "conversion_rate": round(conversion_rate, 1)
        },
        "top_customers": top_customers,
        "monthly_breakdown": monthly_breakdown
    }

# ============== ADMIN MODELS ==============

class CouponCreate(BaseModel):
    code: str
    discount_type: str  # percent, amount
    discount_value: float
    max_uses: Optional[int] = None
    expires_at: Optional[str] = None
    is_active: bool = True

class CouponResponse(BaseModel):
    id: str
    code: str
    discount_type: str
    discount_value: float
    max_uses: Optional[int] = None
    used_count: int = 0
    expires_at: Optional[str] = None
    is_active: bool
    created_at: str

class CampaignCreate(BaseModel):
    subject: str
    content: str
    recipient_type: str  # all, active, trial, expired

class AdminUserResponse(BaseModel):
    id: str
    email: str
    company_name: str
    subscription_status: str
    trial_end_date: Optional[str] = None
    created_at: str
    quotes_count: int = 0

# ============== ADMIN HELPER ==============

async def check_admin(user: dict):
    """Check if user is admin"""
    if not user.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Yönetici yetkisi gerekli")

# ============== ADMIN ENDPOINTS ==============

@api_router.get("/admin/users")
async def admin_get_users(current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Get quote counts for each user
    for user in users:
        count = await db.quotes.count_documents({"user_id": user["id"]})
        user["quotes_count"] = count
    
    return users

@api_router.get("/admin/stats")
async def admin_get_stats(current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    
    total_users = await db.users.count_documents({})
    active_users = await db.users.count_documents({"subscription_status": "active"})
    trial_users = await db.users.count_documents({"subscription_status": "trial"})
    expired_users = await db.users.count_documents({"subscription_status": "expired"})
    total_quotes = await db.quotes.count_documents({})
    total_coupons = await db.coupons.count_documents({})
    
    return {
        "total_users": total_users,
        "active_users": active_users,
        "trial_users": trial_users,
        "expired_users": expired_users,
        "total_quotes": total_quotes,
        "total_coupons": total_coupons
    }

# ============== COUPON ENDPOINTS ==============

@api_router.get("/admin/coupons")
async def admin_get_coupons(current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    coupons = await db.coupons.find({}, {"_id": 0}).to_list(100)
    return coupons

@api_router.post("/admin/coupons")
async def admin_create_coupon(coupon_data: CouponCreate, current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    
    # Check if code exists
    existing = await db.coupons.find_one({"code": coupon_data.code.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="Bu kupon kodu zaten mevcut")
    
    coupon_doc = {
        "id": str(uuid.uuid4()),
        "code": coupon_data.code.upper(),
        "discount_type": coupon_data.discount_type,
        "discount_value": coupon_data.discount_value,
        "max_uses": coupon_data.max_uses,
        "used_count": 0,
        "expires_at": coupon_data.expires_at,
        "is_active": coupon_data.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.coupons.insert_one(coupon_doc)
    del coupon_doc["_id"]
    return coupon_doc

@api_router.delete("/admin/coupons/{coupon_id}")
async def admin_delete_coupon(coupon_id: str, current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    result = await db.coupons.delete_one({"id": coupon_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kupon bulunamadı")
    return {"message": "Kupon silindi"}

@api_router.put("/admin/coupons/{coupon_id}/toggle")
async def admin_toggle_coupon(coupon_id: str, current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    coupon = await db.coupons.find_one({"id": coupon_id})
    if not coupon:
        raise HTTPException(status_code=404, detail="Kupon bulunamadı")
    
    new_status = not coupon.get("is_active", True)
    await db.coupons.update_one({"id": coupon_id}, {"$set": {"is_active": new_status}})
    return {"message": "Kupon durumu güncellendi", "is_active": new_status}

# Public endpoint for validating coupon
@api_router.post("/coupons/validate")
async def validate_coupon(code: str):
    coupon = await db.coupons.find_one({"code": code.upper(), "is_active": True}, {"_id": 0})
    if not coupon:
        raise HTTPException(status_code=404, detail="Geçersiz kupon kodu")
    
    # Check if expired
    if coupon.get("expires_at"):
        expires = datetime.fromisoformat(coupon["expires_at"].replace("Z", "+00:00"))
        if expires < datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="Kupon süresi dolmuş")
    
    # Check max uses
    if coupon.get("max_uses") and coupon.get("used_count", 0) >= coupon["max_uses"]:
        raise HTTPException(status_code=400, detail="Kupon kullanım limiti dolmuş")
    
    return {
        "valid": True,
        "discount_type": coupon["discount_type"],
        "discount_value": coupon["discount_value"]
    }

# ============== EMAIL CAMPAIGN ENDPOINTS ==============

def send_smtp_email(to_email: str, subject: str, html_content: str):
    """Send email via SMTP"""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = SMTP_USER
        msg['To'] = to_email
        
        html_part = MIMEText(html_content, 'html')
        msg.attach(html_part)
        
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_USER, to_email, msg.as_string())
        
        return True
    except Exception as e:
        logger.error(f"Failed to send email to {to_email}: {e}")
        return False

@api_router.post("/admin/campaigns/send")
async def admin_send_campaign(campaign: CampaignCreate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    
    # Get recipients based on type
    query = {}
    if campaign.recipient_type == "active":
        query["subscription_status"] = "active"
    elif campaign.recipient_type == "trial":
        query["subscription_status"] = "trial"
    elif campaign.recipient_type == "expired":
        query["subscription_status"] = "expired"
    
    users = await db.users.find(query, {"email": 1}).to_list(10000)
    emails = [u["email"] for u in users]
    
    if not emails:
        raise HTTPException(status_code=400, detail="Hiç alıcı bulunamadı")
    
    # Save campaign record
    campaign_doc = {
        "id": str(uuid.uuid4()),
        "subject": campaign.subject,
        "content": campaign.content,
        "recipient_type": campaign.recipient_type,
        "recipient_count": len(emails),
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "sent_by": current_user["id"]
    }
    await db.campaigns.insert_one(campaign_doc)
    
    # Send emails in background
    html_template = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: #0F172A; padding: 20px; text-align: center;">
            <h1 style="color: #F97316; margin: 0;">TeklifMaster</h1>
        </div>
        <div style="padding: 30px; background: #f8fafc;">
            {campaign.content}
        </div>
        <div style="padding: 20px; text-align: center; background: #e2e8f0;">
            <p style="color: #64748b; font-size: 12px;">Bu e-posta TeklifMaster tarafından gönderilmiştir.</p>
        </div>
    </div>
    """
    
    # Send to each recipient
    success_count = 0
    for email in emails:
        if send_smtp_email(email, campaign.subject, html_template):
            success_count += 1
    
    return {
        "message": f"Kampanya {success_count}/{len(emails)} alıcıya gönderildi",
        "total": len(emails),
        "sent": success_count
    }

@api_router.get("/admin/campaigns")
async def admin_get_campaigns(current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    campaigns = await db.campaigns.find({}, {"_id": 0}).sort("sent_at", -1).to_list(100)
    return campaigns

# ============== MAKE USER ADMIN ==============

@api_router.post("/admin/make-admin/{user_email}")
async def make_user_admin(user_email: str, current_user: dict = Depends(get_auth_user)):
    await check_admin(current_user)
    
    result = await db.users.update_one(
        {"email": user_email},
        {"$set": {"is_admin": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    
    return {"message": f"{user_email} artık yönetici"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
